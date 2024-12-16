# gui.py

import json
import multiprocessing
import os
import time
from tkinter import filedialog, messagebox, StringVar

import customtkinter as ctk
import pymupdf
from openpyxl import Workbook, load_workbook

from constants import *
from pdf_processor import PDFProcessor
from pdf_viewer import PDFViewer
from utils import create_tooltip, EditableTreeview
from functools import partial

class ReidactorGUI:
    def __init__(self, root):
        self.root = root
        self.pdf_viewer = PDFViewer(self, self.root)  # Pass GUI instance and root window

        self.include_subfolders = False
        self.pdf_folder = ''
        self.output_excel_path = ''

        self.recent_pdf_path = None

        self.setup_widgets()
        self.setup_bindings()
        self.setup_tooltips()

    def setup_widgets(self):
        # PDF Folder Entry
        self.pdf_folder_entry = ctk.CTkEntry(self.root, width=270, height=20, font=(BUTTON_FONT, 9),
                                             placeholder_text="Select Folder with PDFs", border_width=1,
                                             corner_radius=3)
        self.pdf_folder_entry.place(x=50, y=10)
        self.pdf_folder_button = ctk.CTkButton(self.root, text="...", command=self.browse_pdf_folder,
                                               font=(BUTTON_FONT, 9),
                                               width=25, height=10)
        self.pdf_folder_button.place(x=20, y=10)


        # Zoom Slider
        self.zoom_var = ctk.DoubleVar(value=self.pdf_viewer.current_zoom)  # Initialize with the current zoom level
        self.zoom_slider = ctk.CTkSlider(self.root, from_=0.1, to=4, variable=self.zoom_var,
                                         command=self.update_zoom, width=155)
        self.zoom_slider.place(x=20, y=110)

        # Open Sample PDF Button
        self.open_sample_button = ctk.CTkButton(self.root, text="Open PDF", command=self.open_sample_pdf,
                                                font=(BUTTON_FONT, 9),
                                                width=25, height=10)
        self.open_sample_button.place(x=20, y=35)

        # Recent PDF Button
        self.recent_pdf_button = ctk.CTkButton(self.root, text="Recent", command=self.open_recent_pdf,
                                               font=(BUTTON_FONT, 9), width=40, height=10)
        self.recent_pdf_button.place(x=87, y=35)

        # Close PDF Button
        self.close_pdf_button = ctk.CTkButton(self.root, text="x", command=self.close_pdf,
                                              font=(BUTTON_FONT, 9), width=10, height=10, fg_color="red2")
        self.close_pdf_button.place(x=143, y=35)

        # Output Excel Path
        self.output_path_entry = ctk.CTkEntry(self.root, width=270, height=20, font=(BUTTON_FONT, 9),
                                              placeholder_text="Select Folder for Excel output",
                                              border_width=1, corner_radius=3)
        self.output_path_entry.place(x=50, y=60)
        self.output_path_button = ctk.CTkButton(self.root, text="...", command=self.browse_output_path,
                                                font=(BUTTON_FONT, 9),
                                                width=25, height=10)
        self.output_path_button.place(x=20, y=60)

        # Include Subfolders Checkbox
        self.include_subfolders_var = ctk.IntVar()
        self.include_subfolders_checkbox = ctk.CTkCheckBox(self.root, text="Include Subfolders?",
                                                           variable=self.include_subfolders_var,
                                                           command=self.toggle_include_subfolders,
                                                           font=(BUTTON_FONT, 9),checkbox_width=17, checkbox_height=17)
        self.include_subfolders_checkbox.place(x=192, y=34)


        # Areas Treeview setup
        self.areas_frame = ctk.CTkFrame(self.root, height=1, width=200, border_width=0)
        self.areas_frame.place(x=-425, y=-10)

        self.areas_tree = EditableTreeview(
            self,
            self.areas_frame,
            columns=("Title", "x0", "y0", "x1", "y1"),  # Ensure "Title" is included
            show="headings",
            height=3
        )

        # Set up static headers and fixed column widths
        self.areas_tree.heading("Title", text="Title")
        self.areas_tree.column("Title", width=50, anchor="center")
        for col in ("x0", "y0", "x1", "y1"):
            self.areas_tree.heading(col, text=col)
            self.areas_tree.column(col, width=45, anchor="center")

        # Pack the Treeview into the frame
        self.areas_tree.pack(side="left")

        # Import, Export, and Clear Areas Buttons
        self.import_button = ctk.CTkButton(self.root, text="Import", command=self.import_from_excel,
                                           font=(BUTTON_FONT, 8.5), width=55, height=7)
        self.import_button.place(x=890, y=33)

        self.export_button = ctk.CTkButton(self.root, text="Export", command=self.export_to_excel,
                                           font=(BUTTON_FONT, 8.5), width=55, height=7)
        self.export_button.place(x=890, y=53)

        self.clear_areas_button = ctk.CTkButton(self.root, text="Clear All", command=self.clear_all_areas,
                                                font=(BUTTON_FONT, 8.5), width=55, height=7)
        self.clear_areas_button.place(x=890, y=73)

        # Modes
        self.mode_label = ctk.CTkLabel(self.root, text="MODE:", font=(BUTTON_FONT, 9))
        self.mode_label.place(x=337, y=0)

        self.text_mode_button = ctk.CTkButton(self.root, text="Text", command=self.toggle_text_mode,
                                              font=(BUTTON_FONT, 10), width=80, height=25)
        self.text_mode_button.place(x=330, y=24)

        self.deletion_mode_button = ctk.CTkButton(self.root, text="Redact", command=self.toggle_deletion_mode,
                                                  font=(BUTTON_FONT, 10), width=80, height=25)
        self.deletion_mode_button.place(x=330, y=54)

        # Font Style Dropdown
        self.font_label = ctk.CTkLabel(self.root, text="Font:", font=(BUTTON_FONT, 9))
        self.font_label.place(x=420, y=13)

        self.font_style_var = StringVar(value="Helvetica")  # Default font
        self.font_styles = list(pymupdf.Base14_fontdict.values())  # Base-14 font styles
        self.font_style_menu = ctk.CTkOptionMenu(
            self.root, dynamic_resizing=False,
            values=self.font_styles,
            variable=self.font_style_var,
            font=(BUTTON_FONT, 9),
            width=88,
            height=18
        )
        self.font_style_menu.place(x=420, y=35)  # Adjust position as needed



        # Font Size Textbox
        self.font_size_var = StringVar(value="9")  # Default font size
        self.font_size_entry = ctk.CTkEntry(
            self.root,
            textvariable=self.font_size_var,
            font=(BUTTON_FONT, 9),
            width=40,
            height=20
        )
        self.font_size_entry.place(x=420, y=56)  # Adjust position as needed

        # self.size_label = ctk.CTkLabel(self.root, text="Font Size:", font=(BUTTON_FONT, 9))
        # self.size_label.place(x=423, y=38)

        # Revision Updater Checkbox
        self.revision_updater_var = ctk.IntVar()
        self.revision_updater_checkbox = ctk.CTkCheckBox(
            self.root, text="Revision Updater",
            variable=self.revision_updater_var,
            command=self.toggle_revision_updater,
            font=(BUTTON_FONT, 9),
            checkbox_width=17, checkbox_height=17
        )
        self.revision_updater_checkbox.place(x=522, y=5)  # Adjust position as needed

        self.clipping_area_button = ctk.CTkButton(
            self.root, text="History Area",
            font=(BUTTON_FONT, 9), width=60, height=25,
            command=self.set_clipping_area
        )
        self.clipping_area_button.place(x=522, y=30)  # Adjust position as needed
        self.clipping_area_button.configure(state="disabled")  # Disable by default

        self.rev_area_button = ctk.CTkButton(
            self.root, text="Rev No Area",
            font=(BUTTON_FONT, 9), width=60, height=25,
            command=self.set_revision_area
        )
        self.rev_area_button.place(x=522, y=60)  # Adjust position as needed
        self.rev_area_button.configure(state="disabled")  # Disable by default



        # Date Entry
        self.date_entry = ctk.CTkEntry(
            self.root, width=150, height=20, border_width=1, corner_radius=3,
            font=(BUTTON_FONT, 9), placeholder_text="Rev Date (e.g., 09-Jan-25)"
        )
        self.date_entry.place(x=610, y=30)  # Adjust position as needed
        self.date_entry.configure(state="disabled")  # Disable by default

        # Description Entry
        self.description_entry = ctk.CTkEntry(
            self.root, width=150, height=20, border_width=1, corner_radius=3,
            font=(BUTTON_FONT, 9), placeholder_text="Rev Description (e.g., Issued for Tender)"
        )
        self.description_entry.place(x=610, y=60)  # Adjust position as needed
        self.description_entry.configure(state="disabled")  # Disable by default

        # Process Button
        self.extract_button = ctk.CTkButton(self.root, text="PROCESS", font=("Arial Black", 12),
                                            corner_radius=10, width=75, height=55, command=self.start_processing)
        self.extract_button.place(x=793, y=35)


        # Version Label with Tooltip
        self.version_label = ctk.CTkLabel(self.root, text=VERSION_TEXT, fg_color="transparent",
                                          text_color="gray59",
                                          font=(BUTTON_FONT, 9.5))
        self.version_label.place(x=835, y=5)
        self.version_label.bind("<Button-1>", self.display_version_info)

    def export_rectangles(self):
        """Exports the currently selected areas (rectangles) to a JSON file."""
        export_file_path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Save Rectangles As"
        )
        if export_file_path:
            try:
                with open(export_file_path, 'w', encoding='utf-8') as json_file:
                    json.dump(self.pdf_viewer.areas, json_file, indent=4)
                print(f"Exported areas to {export_file_path}")
            except Exception as e:
                messagebox.showerror("Export Error", f"Could not export areas: {e}")

    def import_rectangles(self):
        """Imports area selections from a JSON file."""
        import_file_path = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="Import Rectangles"
        )
        if import_file_path:
            try:
                with open(import_file_path, 'r') as json_file:
                    imported_areas = json.load(json_file)
                self.pdf_viewer.areas = imported_areas  # Update the areas in the PDF viewer
                self.pdf_viewer.update_rectangles()  # Refresh the rectangles on the canvas
                self.update_areas_treeview()  # Refresh the Treeview to show imported areas
                print(f"Imported areas from {import_file_path}")
            except Exception as e:
                messagebox.showerror("Import Error", f"Could not import areas: {e}")

    def clear_all_areas(self):
        """Clears all areas and updates the display."""
        self.pdf_viewer.clear_areas()  # Clear areas from the PDF viewer
        self.areas_tree.delete(*self.areas_tree.get_children())  # Clear all entries in the Treeview
        print("All areas cleared.")

    def export_to_excel(self):
        """Exports deletion areas, insertion points, and table/revision coordinates to an Excel file."""
        export_file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Export Areas"
        )

        if not export_file_path:
            return  # User canceled the save dialog

        try:
            # Create a new Excel workbook
            wb = Workbook()

            # Add Deletion Areas sheet
            ws_deletion = wb.active
            ws_deletion.title = "Deletion Areas"
            ws_deletion.append(["X0", "Y0", "X1", "Y1", "Title"])  # Headers
            for area in self.pdf_viewer.areas:
                coordinates = area["coordinates"]
                title = area.get("title", "Untitled")
                ws_deletion.append(coordinates + [title])

            # Add Insertion Points sheet
            ws_insertion = wb.create_sheet(title="Insertion Points")
            ws_insertion.append(["X", "Y", "Text", "Font", "Size"])  # Headers
            for point in self.pdf_viewer.insertion_points:
                ws_insertion.append([
                    point["position"][0],
                    point["position"][1],
                    point["text"],
                    point["font"],
                    point["size"]
                ])

            # Add Table and Revision Areas sheet
            ws_table_revision = wb.create_sheet(title="Table and Revision Areas")
            ws_table_revision.append(["Type", "X0", "Y0", "X1", "Y1"])  # Headers
            if self.pdf_viewer.table_coordinates:
                ws_table_revision.append(["Table"] + self.pdf_viewer.table_coordinates)
            if self.pdf_viewer.rev_coordinates:
                ws_table_revision.append(["Revision"] + self.pdf_viewer.rev_coordinates)

            # Save the workbook
            wb.save(export_file_path)
            print(f"Exported to Excel at {export_file_path}")
            messagebox.showinfo("Export Successful",
                                "Areas, insertion points, and coordinates have been exported to Excel.")
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            messagebox.showerror("Export Error", f"An error occurred while exporting to Excel: {e}")

    def import_from_excel(self):
        """Imports deletion areas, insertion points, and table/revision coordinates from an Excel file."""
        import_file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Import Areas"
        )

        if not import_file_path:
            return  # User canceled the open dialog

        try:
            # Load the Excel workbook
            wb = load_workbook(import_file_path)

            # Read Deletion Areas
            if "Deletion Areas" in wb.sheetnames:
                ws_deletion = wb["Deletion Areas"]
                self.pdf_viewer.areas = []  # Clear existing deletion areas

                for row in ws_deletion.iter_rows(min_row=2, values_only=True):  # Skip the header row
                    x0, y0, x1, y1, title = row
                    self.pdf_viewer.areas.append({
                        "coordinates": [x0, y0, x1, y1],
                        "title": title
                    })

            # Read Insertion Points
            if "Insertion Points" in wb.sheetnames:
                ws_insertion = wb["Insertion Points"]
                self.pdf_viewer.insertion_points = []  # Clear existing insertion points

                for row in ws_insertion.iter_rows(min_row=2, values_only=True):  # Skip the header row
                    x, y, text, font, size = row
                    self.pdf_viewer.insertion_points.append({
                        "position": [x, y],
                        "text": text,
                        "font": font,
                        "size": int(size) if size else 12  # Default size if missing
                    })

            # Read Table and Revision Areas
            if "Table and Revision Areas" in wb.sheetnames:
                ws_table_revision = wb["Table and Revision Areas"]

                # Check if the sheet has at least one row of data
                rows = list(ws_table_revision.iter_rows(min_row=2, values_only=True))  # Skip the header row
                if rows:
                    for row in rows:
                        if any(row):  # Skip empty rows
                            area_type, x0, y0, x1, y1 = row
                            if area_type == "Table":
                                self.pdf_viewer.table_coordinates = [x0, y0, x1, y1]
                            elif area_type == "Revision":
                                self.pdf_viewer.rev_coordinates = [x0, y0, x1, y1]
                else:
                    print("Table and Revision Areas sheet is blank. Skipping update.")

            # Refresh the canvas and Treeview
            self.pdf_viewer.update_rectangles()
            self.update_areas_treeview()

            print(f"Imported areas, insertion points, and coordinates from {import_file_path}")
            messagebox.showinfo("Import Successful",
                                "Areas, insertion points, and coordinates have been imported successfully.")
        except Exception as e:
            print(f"Error importing from Excel: {e}")
            messagebox.showerror("Import Error", f"An error occurred while importing from Excel: {e}")

    def update_areas_treeview(self):
        """Refreshes the Treeview to display the current areas and their titles."""

        # Clear existing entries
        self.areas_tree.delete(*self.areas_tree.get_children())

        # Insert each area with its title and coordinates into the Treeview and keep track of each item ID
        self.treeview_item_ids = {}  # Dictionary to track Treeview item IDs to canvas rectangle IDs

        for index, area in enumerate(self.pdf_viewer.areas):
            coordinates = area["coordinates"]
            title = area["title"]
            # Insert row into the Treeview and get its item ID
            item_id = self.areas_tree.insert("", "end", values=(title, *coordinates))
            # Store the item ID associated with the canvas rectangle index
            self.treeview_item_ids[item_id] = index

    def open_sample_pdf(self):
        # Opens a file dialog to select a PDF file, then displays it in the PDFViewer
        pdf_path = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
        if pdf_path:
            self.pdf_viewer.display_pdf(pdf_path)
            self.recent_pdf_path = pdf_path  # Store the recent PDF path
            print(f"Opened sample PDF: {pdf_path}")

    def open_recent_pdf(self):
        """Opens the most recently viewed PDF."""
        if self.recent_pdf_path:
            self.pdf_viewer.display_pdf(self.recent_pdf_path)
            print(f"Reopened recent PDF: {self.recent_pdf_path}")
        else:
            messagebox.showinfo("Info", "No recent PDF found.")

    def close_pdf(self):
        """Delegates PDF closing to the PDF viewer."""
        self.pdf_viewer.close_pdf()

    def remove_row(self):
        """Removes the selected row from the Treeview and updates the canvas to remove the associated rectangle."""
        selected_item = self.areas_tree.selection()
        if selected_item:
            # Get the rectangle index associated with the selected Treeview item
            index = self.treeview_item_ids.get(selected_item[0])
            if index is not None:
                # Remove the rectangle from the canvas
                rectangle_id = self.pdf_viewer.rectangle_list[index]
                self.pdf_viewer.canvas.delete(rectangle_id)
                # Remove the area from PDFViewer's areas and rectangle_list
                del self.pdf_viewer.areas[index]
                del self.pdf_viewer.rectangle_list[index]

                # Remove the item from Treeview
                self.areas_tree.delete(selected_item[0])

                # Update Treeview and canvas display
                self.update_areas_treeview()
                self.pdf_viewer.update_rectangles()

                print("Removed rectangle at index", index)

    def setup_bindings(self):
        self.pdf_folder_entry.bind("<KeyRelease>", self.update_pdf_folder)
        self.output_path_entry.bind("<KeyRelease>", self.update_output_path)
        self.root.bind("<Configure>", self.on_window_resize)

    def toggle_revision_updater(self):
        """Enables or disables the revision updater mode."""
        is_enabled = self.revision_updater_var.get() == 1
        state = "normal" if is_enabled else "disabled"

        self.clipping_area_button.configure(state=state)
        self.rev_area_button.configure(state=state)
        self.date_entry.configure(state=state)  # Enable/disable Date entry
        self.description_entry.configure(state=state)

    def toggle_text_mode(self):
        self.pdf_viewer.set_mode(TEXT_MODE)

        # Update button states
        self.text_mode_button.configure(fg_color="green4", text_color="white")
        self.deletion_mode_button.configure(fg_color="cornsilk4", text_color="black")
        self.clipping_area_button.configure(fg_color="cornsilk4", text_color="black")
        self.rev_area_button.configure(fg_color="cornsilk4", text_color="black")

    def toggle_deletion_mode(self):
        self.pdf_viewer.set_mode(REDACTION_MODE)

        # Update button states
        self.text_mode_button.configure(fg_color="cornsilk4", text_color="black")
        self.deletion_mode_button.configure(fg_color="green3", text_color="white")
        self.clipping_area_button.configure(fg_color="cornsilk4", text_color="black")
        self.rev_area_button.configure(fg_color="cornsilk4", text_color="black")

    def set_clipping_area(self):
        self.pdf_viewer.set_mode(TABLE_COORDINATES_MODE)

        # Update button states
        self.text_mode_button.configure(fg_color="cornsilk4", text_color="black")
        self.deletion_mode_button.configure(fg_color="cornsilk4", text_color="black")
        self.clipping_area_button.configure(fg_color="green4", text_color="white")
        self.rev_area_button.configure(fg_color="cornsilk4", text_color="black")

    def set_revision_area(self):
        self.pdf_viewer.set_mode(REVISION_COORDINATES_MODE)

        # Update button states
        self.text_mode_button.configure(fg_color="cornsilk4", text_color="black")
        self.deletion_mode_button.configure(fg_color="cornsilk4", text_color="black")
        self.clipping_area_button.configure(fg_color="cornsilk4", text_color="black")
        self.rev_area_button.configure(fg_color="green4", text_color="white")

    def setup_tooltips(self):
        create_tooltip(self.pdf_folder_entry, "Select the main folder containing PDF files")
        create_tooltip(self.open_sample_button, "Open a sample PDF to set areas")
        create_tooltip(self.output_path_entry, "Select folder for the Excel output")
        create_tooltip(self.include_subfolders_checkbox, "Include files from subfolders for extraction")
        create_tooltip(self.extract_button, "Start the extraction process")
        create_tooltip(self.import_button, "Import a saved template of selected areas")
        create_tooltip(self.export_button, "Export the selected areas as a template")
        create_tooltip(self.clear_areas_button, "Clear all selected areas")

    def browse_pdf_folder(self):
        self.pdf_folder = filedialog.askdirectory()
        self.pdf_folder_entry.delete(0, ctk.END)
        self.pdf_folder_entry.insert(0, self.pdf_folder)

    def browse_output_path(self):
        """Opens a dialog to specify the output folder."""
        selected_folder = filedialog.askdirectory()
        if selected_folder:  # Only update the entry if a folder was selected
            self.output_excel_path = selected_folder  # Store the selected folder path
            self.output_path_entry.delete(0, ctk.END)
            self.output_path_entry.insert(0, self.output_excel_path)

    def update_zoom_slider(self, zoom_level):
        """Updates the zoom slider to reflect the current zoom level in PDFViewer."""
        self.zoom_var.set(zoom_level)

    def update_pdf_folder(self, event):
        self.pdf_folder = self.pdf_folder_entry.get()

    def update_output_path(self, event):
        self.output_excel_path = self.output_path_entry.get()

    def update_zoom(self, value):
        """Adjusts the zoom level of the PDFViewer based on slider input."""
        zoom_level = float(value)
        self.pdf_viewer.set_zoom(zoom_level)  # Update zoom in PDFViewer

    def toggle_include_subfolders(self):
        self.include_subfolders = self.include_subfolders_var.get()

    def start_processing(self):
        self.start_time = time.time()

        self.pdf_viewer.close_pdf()  # Ensure no open files interfere

        # Setup progress window
        self.progress_window = ctk.CTkToplevel(self.root)
        self.progress_window.title("Progress")
        self.progress_window.geometry("300x120")

        # Add a progress label
        self.progress_label = ctk.CTkLabel(self.progress_window, text="Processing PDFs...")
        self.progress_label.pack(pady=5)

        # Add a total files label
        self.total_files_label = ctk.CTkLabel(self.progress_window, text="Total files: 0")
        self.total_files_label.pack(pady=5)

        self.progress_var = ctk.DoubleVar(value=0)
        self.progress_bar = ctk.CTkProgressBar(self.progress_window, variable=self.progress_var,
                                               orientation="horizontal", width=250)
        self.progress_bar.pack(pady=10)

        # Fetch the Date and Description values from the text boxes
        date_value = self.date_entry.get()
        description_value = self.description_entry.get()

        # Validation: Ensure values are provided
        if not date_value or not description_value:
            messagebox.showerror("Missing Information",
                                 "Please provide both Date and Description for the revision updater.")
            return

        manager = multiprocessing.Manager()
        progress_list = manager.list()
        total_files = manager.Value('i', 0)

        processor1 = PDFProcessor(
            pdf_folder=self.pdf_folder,
            output_excel_path=self.output_excel_path,
            areas=self.pdf_viewer.areas,
            insertion_points=self.pdf_viewer.insertion_points,
            include_subfolders=self.include_subfolders,
            table_coordinates = self.pdf_viewer.table_coordinates,
            rev_coordinates = self.pdf_viewer.rev_coordinates,
            revision_date=date_value,
            revision_description=description_value
        )

        pdf_files = processor1.get_pdf_files()
        total_files.value = len(pdf_files)

        if not pdf_files:
            messagebox.showinfo("No Files", "No PDF files found in the selected folder.")
            self.progress_window.destroy()
            return

        pool = multiprocessing.Pool()
        process_func = partial(processor1.process_single_pdf, progress_list=progress_list)
        pool.map_async(process_func, pdf_files)

        self.root.after(100, self.update_progress, progress_list, total_files, pool)

    def update_progress(self, progress_list, total_files, pool):
        """Updates the progress bar based on the progress of PDF extraction."""
        try:
            # Avoid division by zero
            if total_files.value > 0:
                current_progress = len(progress_list) / total_files.value
                self.progress_var.set(current_progress)

                # Update the progress label with the number of files processed
                progress_text = f"Processed {len(progress_list)} of {total_files.value} files."
                self.total_files_label.configure(text=progress_text)  # Update the title or add a label to display progress

            # Check if all tasks in the pool are complete
            if not pool._cache:  # Pool tasks are done when the cache is empty
                self.progress_var.set(1)  # Ensure progress bar is complete
                self.progress_window.destroy()  # Close progress window

                # Calculate and display elapsed time
                end_time = time.time()
                elapsed_time = end_time - self.start_time
                formatted_time = time.strftime("%H:%M:%S", time.gmtime(elapsed_time))

                response = messagebox.askyesno(
                    "Processing Complete",
                    f"PDF processing completed successfully in {formatted_time}.\nWould you like to open the output folder?"
                )
                if response:
                    os.startfile(self.output_excel_path)
            else:
                # Continue monitoring progress
                self.root.after(100, self.update_progress, progress_list, total_files, pool)
        except Exception as e:
            print(f"Error updating progress: {e}")

    def on_window_resize(self, event):
        """Handles window resizing and adjusts the canvas dimensions."""
        self.pdf_viewer.resize_canvas()

    def display_version_info(self, event):
        version_text = """
        Created by: Rei Raphael Reveral

        Links:
        https://github.com/r-Yayap/MultiplePDF-Areas2Excel
        https://www.linkedin.com/in/rei-raphael-reveral
        """
        window = ctk.CTkToplevel(self.root)
        window.title("Version Info")
        text_widget = ctk.CTkTextbox(window, wrap="word", width=400, height=247)
        text_widget.insert("end", version_text)
        text_widget.pack(padx=10, pady=10, side="left")
        window.grab_set()

